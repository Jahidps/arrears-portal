#!/usr/bin/env python3
"""
Converts the weekly arrears xlsx files into data.json + data.js for the portal.

data/ folder:
  - "Del Report (DD-MM-YYYY).xlsx"           -> FDGL / Paytek tabs
  - "PS Lease Del report -DD-MM-YYYY.xlsx"   -> PS Lease tab

The portal always shows the NEWEST report of each type. But ALL reports in
data/ are processed in date order to build the cumulative "Settled" tab:
any lease present in one week's report but missing from the next is settled
(and removed again if it ever reappears). Previously settled leases from the
existing data.json are carried forward, so deleting old xlsx files is safe.

Run:  python scripts/convert.py
"""
import json
import re
import sys
from datetime import datetime, date, time
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent.parent
DATA_DIR = ROOT / "data"
OUT_FILE = ROOT / "data.json"

DATE_RE = re.compile(r"(\d{2})-(\d{2})-(\d{4})")
MONTH_RE = re.compile(r"(january|february|march|april|may|june|july|august|september|october|november|december)[ _-]*(\d{4})", re.IGNORECASE)
MONTHS = {m: i+1 for i, m in enumerate(
    ["january","february","march","april","may","june",
     "july","august","september","october","november","december"])}

SETTLED_COLS = ["Provider", "MID", "Legal Name", "Trading Name", "Lease No.",
                "Last Arrears", "Paid", "Last Status", "Settled On"]


def file_date(path: Path):
    m = DATE_RE.search(path.name)
    if m:
        d, mo, y = map(int, m.groups())
        try:
            return datetime(y, mo, d)
        except ValueError:
            pass
    m = MONTH_RE.search(path.name)
    if m:
        return datetime(int(m.group(2)), MONTHS[m.group(1).lower()], 1)
    return datetime.fromtimestamp(path.stat().st_mtime)


def files_sorted(pattern: str):
    files = [p for p in DATA_DIR.glob("*.xlsx")
             if re.match(pattern, p.name, re.IGNORECASE) and not p.name.startswith("~")]
    return sorted(files, key=file_date)


def clean(v):
    if v is None:
        return ""
    if isinstance(v, datetime):
        return v.strftime("%d/%m/%Y")
    if isinstance(v, date):
        return v.strftime("%d/%m/%Y")
    if isinstance(v, time):
        return ""  # a time-only value in a date column is not usable
    if isinstance(v, float) and v == int(v):
        return int(v)
    if isinstance(v, str):
        return v.strip()
    if isinstance(v, (int, float, bool)):
        return v
    return str(v)


def read_sheet(ws):
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return [], []
    header = [str(h).strip() if h is not None else "" for h in rows[0]]
    while header and header[-1] == "":
        header.pop()
    data = []
    for r in rows[1:]:
        vals = [clean(v) for v in r[: len(header)]]
        if any(v != "" for v in vals):
            data.append(vals)
    return header, data


def col(columns, *names):
    low = [c.lower() for c in columns]
    for n in names:
        if n in low:
            return low.index(n)
    return -1


def lease_keys(tab):
    i = col(tab["columns"], "lease no.", "lease")
    if i < 0:
        return set()
    return {str(r[i]) for r in tab["rows"] if i < len(r)}


def settled_entry(provider, columns, row, settled_on):
    g = lambda *names: (row[col(columns, *names)]
                        if -1 < col(columns, *names) < len(row) else "")
    return [provider,
            g("mid"), g("legal name"), g("trading name"),
            g("lease no.", "lease"), g("arrears"),
            g("paid", "payments made"),
            g("lease status", "mandate status", "back office stage"),
            settled_on]


def build_del_tabs(path: Path):
    """Parse one Del Report file into fdgl/paytek tabs with tag columns."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    header, rows = read_sheet(wb.worksheets[0])
    lease_idx = col(header, "lease no.")
    if lease_idx < 0:
        lease_idx = 3

    def sheet_leases(name):
        if name not in wb.sheetnames:
            return set()
        th, tr = read_sheet(wb[name])
        tl = col(th, "lease no.")
        if tl < 0:
            tl = 3
        return {str(r[tl]) for r in tr if tl < len(r)}

    ps_team = sheet_leases("PS Team")
    new_arr = sheet_leases("New in Arrears")
    wb.close()

    header = header + ["PS Team", "New in Arrears"]
    fdgl, paytek = [], []
    for r in rows:
        lease_raw = str(r[lease_idx]) if lease_idx < len(r) else ""
        r = list(r) + ["Yes" if lease_raw in ps_team else "",
                       "Yes" if lease_raw in new_arr else ""]
        (paytek if "PSAVE" in lease_raw.upper() else fdgl).append(r)
    return {"fdgl": {"label": "FDGL", "columns": header, "rows": fdgl},
            "paytek": {"label": "Paytek", "columns": header, "rows": paytek}}


def build_ps_tab(path: Path):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    h, r = read_sheet(wb.worksheets[0])
    wb.close()
    return {"ps_lease": {"label": "PS Lease", "columns": h, "rows": r}}


def chain(files, builder, settled, seen):
    """Process files oldest->newest; diff consecutive reports into settled.
    Returns (newest_tabs, previous_tabs) — previous is the report before the
    newest one, used for arrears-increase tracking."""
    prev_tabs = None
    before = None
    for f in files:
        tabs_i = builder(f)
        if prev_tabs:
            settled_on = file_date(f).strftime("%d/%m/%Y")
            for k, pt in prev_tabs.items():
                nt = tabs_i.get(k)
                if not nt:
                    continue
                new_keys = lease_keys(nt)
                li = col(pt["columns"], "lease no.", "lease")
                if li < 0:
                    continue
                for r in pt["rows"]:
                    key = str(r[li]) if li < len(r) else ""
                    if key and key not in new_keys and key not in seen:
                        settled.append(settled_entry(pt["label"], pt["columns"], r, settled_on))
                        seen.add(key)
        before = prev_tabs
        prev_tabs = tabs_i
    return prev_tabs, before


def annotate_increase(latest, previous):
    """Add 'Prev Arrears' + 'Increased' columns to each provider tab by
    comparing against the previous report."""
    if not latest:
        return
    for k, nt in latest.items():
        pt = previous.get(k) if previous else None
        li = col(nt["columns"], "lease no.", "lease")
        ai = col(nt["columns"], "arrears")
        prev_map = {}
        if pt:
            pli = col(pt["columns"], "lease no.", "lease")
            pai = col(pt["columns"], "arrears")
            if pli > -1 and pai > -1:
                prev_map = {str(r[pli]): r[pai] for r in pt["rows"]
                            if pli < len(r) and pai < len(r)}
        # place "Prev Arrears" right next to "Arrears"
        ins = ai + 1 if ai > -1 else len(nt["columns"])
        nt["columns"] = nt["columns"][:ins] + ["Prev Arrears"] + nt["columns"][ins:] + ["Increased"]
        for r in nt["rows"]:
            key = str(r[li]) if -1 < li < len(r) else ""
            pv = prev_map.get(key, "")
            inc = ""
            if pv != "" and ai > -1 and ai < len(r):
                try:
                    if float(r[ai]) > float(pv):
                        inc = "Yes"
                except (TypeError, ValueError):
                    pass
            r[ins:ins] = [pv]
            r.append(inc)


def main():
    del_files = files_sorted(r"del report")
    ps_files = files_sorted(r"ps lease")

    if not del_files and not ps_files:
        sys.exit("No xlsx files found in data/ - nothing to do.")

    # carry forward previously settled leases from the existing data.json
    carried = []
    if OUT_FILE.exists():
        try:
            prev = json.loads(OUT_FILE.read_text(encoding="utf-8"))
            carried = prev.get("tabs", {}).get("settled", {}).get("rows", [])
        except Exception:
            carried = []

    settled, seen = [], set()
    li = SETTLED_COLS.index("Lease No.")
    for r in carried:
        key = str(r[li]) if li < len(r) else ""
        if key and key not in seen:
            settled.append(r)
            seen.add(key)

    tabs = {}
    sources = {}

    del_tabs, del_prev = chain(del_files, build_del_tabs, settled, seen)
    if del_tabs:
        annotate_increase(del_tabs, del_prev)
        tabs.update(del_tabs)
        f = del_files[-1]
        print(f"Del Report file : {f.name}  ({len(del_files)} report(s) chained)")
        sources["del_report"] = {"file": f.name,
                                 "date": file_date(f).strftime("%d/%m/%Y")}

    ps_tabs, ps_prev = chain(ps_files, build_ps_tab, settled, seen)
    if ps_tabs:
        annotate_increase(ps_tabs, ps_prev)
        tabs.update(ps_tabs)
        f = ps_files[-1]
        print(f"PS Lease file   : {f.name}  ({len(ps_files)} report(s) chained)")
        sources["ps_lease"] = {"file": f.name,
                               "date": file_date(f).strftime("%d/%m/%Y")}

    # drop settled leases that reappeared in the latest reports
    all_new_keys = set()
    for k in ["fdgl", "paytek", "ps_lease"]:
        if k in tabs:
            all_new_keys |= lease_keys(tabs[k])
    settled = [r for r in settled if str(r[li]) not in all_new_keys]

    tabs["settled"] = {"label": "Settled", "columns": SETTLED_COLS, "rows": settled}

    # ---------- estate report (monthly, searchable only — no tab) ----------
    estate_files = files_sorted(r"estate report")
    if estate_files:
        ef = estate_files[-1]
        print(f"Estate Report   : {ef.name}")
        sources["estate"] = {"file": ef.name,
                             "date": file_date(ef).strftime("%B %Y")}
        wb = openpyxl.load_workbook(ef, read_only=True, data_only=True)
        eh, er = read_sheet(wb.worksheets[0])
        # optional extra sheet holding the PS Lease estate (different columns);
        # match any non-primary sheet whose name mentions PS + estate/lease
        # e.g. "PS Lease Estate", "PS Estate -July"
        pse_h, pse_r = [], []
        main_name = wb.sheetnames[0]
        for sn in wb.sheetnames:
            low = sn.strip().lower()
            if sn != main_name and "ps" in low and ("estate" in low or "lease" in low):
                pse_h, pse_r = read_sheet(wb[sn])
                break
        wb.close()

        eli = col(eh, "lease no", "lease no.", "lease")
        # merge current arrears + status from the del report
        amap = {}
        for k in ["fdgl", "paytek"]:
            t = tabs.get(k)
            if not t:
                continue
            tli = col(t["columns"], "lease no.")
            tai = col(t["columns"], "arrears")
            tsi = col(t["columns"], "lease status")
            for r in t["rows"]:
                amap[str(r[tli])] = (r[tai] if tai > -1 else "",
                                     r[tsi] if tsi > -1 else "")
        eh = eh + ["Arrears", "Lease Status"]
        for r in er:
            key = str(r[eli]) if -1 < eli < len(r) else ""
            a, s = amap.get(key, ("", ""))
            r.extend([a, s])
        tabs["estate"] = {"label": "Estate", "columns": eh, "rows": er}

        # ---- PS Lease estate sheet (searchable only) ----
        # the sheet carries a Yes/No "Arrears" flag; the actual £ amount and
        # missed months come from the PS Lease del report, keyed on Lease No.
        if pse_h:
            # rename the Yes/No flag so it doesn't clash with the £ amount
            pse_h = ["In Arrears" if str(c).strip().lower() == "arrears" else c
                     for c in pse_h]
            pli = col(pse_h, "lease no", "lease no.", "lease")
            pmap = {}
            t = tabs.get("ps_lease")
            if t:
                tli = col(t["columns"], "lease", "lease no.")
                tai = col(t["columns"], "arrears")
                tmm = col(t["columns"], "missed months")
                for r in t["rows"]:
                    pmap[str(r[tli])] = (r[tai] if tai > -1 else "",
                                         r[tmm] if tmm > -1 else "")
            # place the £ "Arrears" amount right after the "In Arrears" flag
            fi = col(pse_h, "in arrears")
            ins = fi + 1 if fi > -1 else len(pse_h)
            pse_h = pse_h[:ins] + ["Arrears"] + pse_h[ins:] + ["Missed Months"]
            for r in pse_r:
                key = str(r[pli]) if -1 < pli < len(r) else ""
                a, mm = pmap.get(key, ("", ""))
                r[ins:ins] = [a]
                r.append(mm)
            tabs["ps_lease_estate"] = {"label": "PS Lease Estate",
                                       "columns": pse_h, "rows": pse_r}
            print(f"PS Lease Estate : {ef.name}  ({len(pse_r)} leases)")

    out = {
        "generated": datetime.now().strftime("%d/%m/%Y %H:%M"),
        "sources": sources,
        "tabs": tabs,
    }
    payload = json.dumps(out, ensure_ascii=False, separators=(",", ":"))
    OUT_FILE.write_text(payload, encoding="utf-8")
    (ROOT / "data.js").write_text("window.PORTAL_DATA=" + payload + ";",
                                  encoding="utf-8")

    for k, t in tabs.items():
        print(f"  {t['label']:<15} {len(t['rows'])} rows")
    print(f"Wrote {OUT_FILE.name} and data.js")


if __name__ == "__main__":
    main()
